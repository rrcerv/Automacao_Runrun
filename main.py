import pandas as pd
import openpyxl
from Excecoes_Customizadas import AutomacaoError
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

nome_output = 'output/output.xlsx'
output_log = 'output/log.txt'


# ---------------- CARREGANDO E PROCESSANDO PLANILHA ------------------------------
comeco = datetime.now()

planilha = 'timesheet-por-cliente-2023-09-12-13h-42m-41s.xlsx'

# TEM QUE TIRAR A PÁGINA DE FILTRO DA PLANILHA SE NÃO ELA NÃO RODA
try:
    customers = pd.read_excel(planilha,
    sheet_name='Timesheet Por Cliente',
    header=0,
    index_col=False,
    keep_default_na=True
    )

    customers['Total de horas'] = customers['Total de horas'].astype('int')

except Exception as e:
    raise AutomacaoError('É necessário abrir a planilha no Excel e remover a aba "Filtro Aplicado".') from None

# --------------------------------- FIM  ----------------------------------------





# -------------------------- GERANDO DATAFRAMES  --------------------------------


# Horas que cada colaborador gastou em um determinado projeto
df_filtrado = customers[['Cliente', 'Projeto', 'Usuário', 'Total de horas']]
horas_por_projeto = df_filtrado.groupby(by=['Cliente', 'Projeto', 'Usuário']).sum().reset_index()


# Total de horas por Cliente
so_horas = customers[['Cliente', 'Total de horas']]
total_horas_clientes = (so_horas.groupby(by='Cliente').sum())



# --------------------------------- FIM  ----------------------------------------



# ----------------------------- GERANDO LOGS  -----------------------------------

log_dic = {}


muitas_horas = customers[customers['Total de horas'] > 8].index.tolist()
for ad in muitas_horas:
    try:
        log_dic[ad] += 'Muitas horas, '
    except:
        log_dic[ad] = 'Muitas horas, '


sem_subgrupo = customers[customers['Subgrupo de Projeto'] == 'Sem subgrupo'].index.tolist()
for ad in sem_subgrupo:
    try:
        log_dic[ad] += 'Sem subgrupo, '
    except:
        log_dic[ad] = 'Sem subgrupo, '


sem_grupo = customers[customers['Grupo de Projeto'] == 'Sem grupo'].index.tolist()
for ad in sem_grupo:
    try:
        log_dic[ad] += 'Sem grupo, '
    except:
        log_dic[ad] = 'Sem grupo, '


sem_cliente = customers[customers['Cliente'] == 'Sem cliente'].index.tolist()
for ad in sem_cliente:
    try:
        log_dic[ad] += 'Sem cliente, '
    except:
        log_dic[ad] = 'Sem cliente, '


sem_projeto = customers[customers['Projeto'] == 'Sem projeto'].index.tolist()
for ad in sem_projeto:
    try:
        log_dic[ad] += 'Sem projeto, '
    except:
        log_dic[ad] = 'Sem projeto, '


sem_tipo = customers[customers['Tipo'] == 'Sem tipo'].index.tolist()
for ad in sem_tipo:
    try:
        log_dic[ad] += 'Sem tipo, '
    except:
        log_dic[ad] = 'Sem tipo, '


df_bat = customers[customers['Cliente'] == 'BAT']
bat_sem_tag = df_bat[df_bat['Tags'].isna()].index.tolist()
for ad in bat_sem_tag:
    try:
        log_dic[ad] += 'BAT sem tag'
    except:
        log_dic[ad] = 'BAT sem tag'


sem_quadro = customers[customers['Quadro'].isna()].index.tolist()
for ad in sem_quadro:
    try:
        log_dic[ad] += 'Sem quadro, '
    except:
        log_dic[ad] = 'Sem quadro, '


with open(output_log, 'w') as f:
    f.write(str(log_dic))

# --------------------------------- FIM  ----------------------------------------



# ----------------------------- FORMAT. SHEET  -----------------------------------

print(log_dic)
lista_ids = list(log_dic.keys())

# GERA UM EXCEL SÓ COM OS REGISTROS QUE CONTÉM ERROS
df_erros = customers.iloc[lista_ids]
df_erros.to_excel(nome_output)

# ABRE O EXCEL E FORMATA COM FUNDO AMARELO AS ENTRADAS QUE FALTAM INFORMAÇÃO
workbook = openpyxl.load_workbook(filename=nome_output)
sheet = workbook.active

#CRIA FORMATAÇÃO PRA TODAS AS CÉLULAS QUE CONTÉM A PALAVRA 'SEM'
formula = ['ISNUMBER(SEARCH("Sem", A1))']
rule = FormulaRule(formula=formula, stopIfTrue=True,
                   fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))
sheet.conditional_formatting.add('A1:U700', rule)

#CRIA FORMATAÇÃO PRA TODAS AS ENTRADAS DA BAT SEM TAG
formula2 = ['AND(U1="BAT", T1="")']
rule2 = FormulaRule(formula=formula2, stopIfTrue=True,
                   fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))
sheet.conditional_formatting.add('T1:T700', rule2)

#CRIA FORMATAÇÃO PRA TODAS AS ENTRADAS QUE TEM HORA MAIOR QUE 8
formula3 = ['AND(Q1>8, Q1<>"Total de horas")']
rule3 = FormulaRule(formula=formula3, stopIfTrue=True,
                    fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))
sheet.conditional_formatting.add('Q1:Q700', rule3)

#CRIA FORMATAÇÃO PRA TODAS AS ENTRADAS QUE NÃO TEM QUADRO
formula4 = ['AND(B1="", A1<>"")']
rule4 = FormulaRule(formula=formula4, stopIfTrue=True,
                    fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))
sheet.conditional_formatting.add('B1:B700', rule4)

workbook.save(nome_output)

# --------------------------------- FIM  ----------------------------------------



fim = datetime.now()
print(f'Tempo de execução do código: {(fim - comeco).total_seconds()} segundos.')
print('Fim.')